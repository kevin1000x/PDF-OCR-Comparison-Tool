"""
PDF智能OCR分类与比对系统 - 主程序
"""

import os
import sys
import argparse
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, field
from datetime import datetime
import yaml

# 导入自定义模块
from project_detector import ProjectDetector, ProjectFolderMapper
from document_classifier import DocumentClassifier, ClassificationResult
from content_matcher import ContentMatcher, PageFeatures, MatchResult, DocumentMatcher
from pdf_processor import PDFProcessor

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def get_ocr_engine(config: dict):
    """
    根据配置获取OCR引擎
    
    Args:
        config: OCR配置字典
        
    Returns:
        OCR引擎实例
    """
    engine_type = config.get('engine', 'paddleocr').lower()
    
    if engine_type == 'deepseek-ocr2':
        logger.info("Using DeepSeek-OCR2 engine")
        from deepseek_ocr2_engine import DeepSeekOCR2Engine
        return DeepSeekOCR2Engine(config)
    else:
        logger.info("Using PaddleOCR engine")
        from ocr_engine import OCREngine
        return OCREngine(config)


def get_ocr_extractor():
    """获取OCR结果提取器"""
    from ocr_engine import OCRResultExtractor
    return OCRResultExtractor()


@dataclass
class ProcessingResult:
    """处理结果"""
    file_path: str
    project: str
    pages: List[dict] = field(default_factory=list)
    matches: List[MatchResult] = field(default_factory=list)
    success: bool = True
    error: str = ""


class DocumentProcessor:
    """文档处理器"""
    
    def __init__(self, config: dict):
        """
        初始化文档处理器
        
        Args:
            config: 配置字典
        """
        self.config = config
        
        # 初始化各组件
        logger.info("Initializing components...")
        
        self.pdf_processor = PDFProcessor(config)
        self.ocr_engine = get_ocr_engine(config.get('ocr', {}))
        self.feature_extractor = get_ocr_extractor()
        self.project_detector = ProjectDetector(config)
        self.classifier = DocumentClassifier(config)
        
        # 进度显示器
        self.progress = None
        self.content_matcher = ContentMatcher(config)
        self.document_matcher = DocumentMatcher(self.content_matcher)
        
        # 初始化OCR缓存
        from ocr_cache import OCRCache
        cache_dir = Path(config.get('paths', {}).get('temp_dir', './temp')) / 'ocr_cache'
        self.ocr_cache = OCRCache(str(cache_dir))
        
        # 路径配置
        paths = config.get('paths', {})
        self.voucher_path = Path(paths.get('input_vouchers', ''))
        self.reference_path = Path(paths.get('reference_docs', ''))
        self.output_path = Path(paths.get('output_dir', ''))
        self.temp_path = Path(paths.get('temp_dir', ''))
        
        # 确保输出目录存在
        self.output_path.mkdir(parents=True, exist_ok=True)
        self.temp_path.mkdir(parents=True, exist_ok=True)
        
        logger.info("Components initialized successfully (with OCR cache)")
        
    def process_pdf_pages(self, pdf_path: str, show_progress: bool = True, use_cache: bool = True, progress_callback=None) -> List[PageFeatures]:
        """
        处理PDF的所有页面，进行OCR和特征提取
        
        Args:
            pdf_path: PDF文件路径
            show_progress: 是否显示页级进度
            use_cache: 是否使用缓存
            progress_callback: 进度回调函数(page_num, total, ocr_time)
            
        Returns:
            页面特征列表
        """
        import time
        
        # 检查缓存
        if use_cache:
            cached = self.ocr_cache.get(pdf_path)
            if cached:
                from ocr_cache import restore_page_features
                page_features = restore_page_features(cached)
                logger.info(f"[CACHED] {pdf_path} ({len(page_features)} pages)")
                return page_features
        
        page_features = []
        
        logger.info(f"Processing PDF: {pdf_path}")
        page_count = self.pdf_processor.get_page_count(pdf_path)
        logger.info(f"Total pages: {page_count}")
        
        start_time = time.time()
        
        for page_num, image in self.pdf_processor.process_pdf(pdf_path):
            page_start = time.time()
            
            # 更新进度
            if progress_callback:
                progress_callback('page_start', page_num, page_count, 0)
            elif show_progress:
                logger.info(f"  OCR page {page_num}/{page_count}...")
            
            # OCR识别
            ocr_result = self.ocr_engine.recognize_pdf_page(image, page_num)
            text = ocr_result.get_full_text()
            
            # 文档类型分类
            classification = self.classifier.classify(text)
            
            # 特征提取
            features = self.feature_extractor.extract_features(text)
            
            # 创建页面特征对象
            page_feature = PageFeatures(
                file_path=pdf_path,
                page_num=page_num,
                text=text,
                doc_type=classification.doc_type,
                dates=features.get('dates', []),
                amounts=features.get('amounts', []),
                numbers=features.get('numbers', []),
                keywords=features.get('keywords', [])
            )
            
            page_features.append(page_feature)
            
            page_time = time.time() - page_start
            
            # 更新进度
            if progress_callback:
                progress_callback('page_done', page_num, page_count, page_time)
            elif show_progress:
                text_preview = text[:50].replace('\n', ' ') + '...' if len(text) > 50 else text.replace('\n', ' ')
                logger.info(f"  Page {page_num} done in {page_time:.1f}s, type={classification.doc_type}, text: {text_preview}")
            
        total_time = time.time() - start_time
        if page_count > 0:
            logger.info(f"PDF processed in {total_time:.1f}s ({page_count} pages, {total_time/page_count:.1f}s/page)")
        
        # 保存到缓存
        if use_cache and page_features:
            self.ocr_cache.put(pdf_path, page_features)
            
        return page_features
        
    def build_reference_index(self, project: Optional[str] = None, use_progress: bool = True):
        """
        构建参照资料索引
        
        Args:
            project: 指定项目名称，None表示构建所有项目
            use_progress: 是否使用进度显示
        """
        import time
        from progress_display import get_progress_display
        
        logger.info("Building reference index...")
        
        mapper = ProjectFolderMapper(self.project_detector, self.config.get('paths', {}))
        reference_files = mapper.scan_all_reference_files()
        
        # 计算总文件数
        total_files = sum(len(files) for name, files in reference_files.items() 
                         if project is None or name == project)
        
        # 初始化进度显示
        progress = get_progress_display(use_progress)
        progress.start(total_files)
        
        all_pages = []
        
        # 进度回调
        def progress_callback(event, page_num, total, ocr_time):
            if event == 'page_start':
                progress.update_page(page_num, total)
            elif event == 'page_done':
                progress.complete_page(ocr_time)
        
        try:
            for proj_name, files in reference_files.items():
                if project and proj_name != project:
                    continue
                
                for pdf_file in files:
                    progress.update_file(str(pdf_file))
                    
                    # 获取页数并更新进度
                    try:
                        page_count = self.pdf_processor.get_page_count(str(pdf_file))
                        progress.add_pages(page_count)
                    except:
                        pass
                    
                    try:
                        pages = self.process_pdf_pages(
                            str(pdf_file), 
                            show_progress=False,
                            progress_callback=progress_callback
                        )
                        all_pages.extend(pages)
                    except Exception as e:
                        logger.error(f"Failed to process: {pdf_file}, error: {e}")
                    
                    progress.complete_file()
        finally:
            progress.stop()
        
        self.content_matcher.build_reference_index(all_pages)
        logger.info(f"Reference index built with {len(all_pages)} pages")
        
    def process_voucher(self, pdf_path: str) -> ProcessingResult:
        """
        处理单个凭证文件
        
        Args:
            pdf_path: PDF文件路径
            
        Returns:
            处理结果
        """
        result = ProcessingResult(file_path=pdf_path, project="")
        
        try:
            # 处理PDF页面
            pages = self.process_pdf_pages(pdf_path)
            
            if not pages:
                result.success = False
                result.error = "No pages processed"
                return result
                
            # 检测项目归属
            all_text = "\n".join([p.text for p in pages])
            project_match = self.project_detector.detect(pdf_path, all_text)
            result.project = project_match.project_name
            
            # 内容匹配
            page_matches = self.content_matcher.match_document(pages)
            merged_matches = self.document_matcher.merge_page_results(page_matches)
            result.matches = merged_matches
            
            # 记录页面信息
            for page in pages:
                result.pages.append({
                    'page_num': page.page_num,
                    'doc_type': page.doc_type,
                    'dates': page.dates,
                    'amounts': page.amounts,
                    'text_preview': page.text[:200] + '...' if len(page.text) > 200 else page.text
                })
                
            # 复制文件到分类目录
            if self.config.get('output', {}).get('copy_files', True):
                # 获取主要文档类型
                doc_types = [p.doc_type for p in pages]
                main_type = max(set(doc_types), key=doc_types.count) if doc_types else "其他"
                
                self.pdf_processor.copy_to_output(
                    pdf_path, 
                    result.project, 
                    main_type
                )
                
            result.success = True
            
        except Exception as e:
            logger.error(f"Failed to process voucher: {pdf_path}, error: {e}")
            result.success = False
            result.error = str(e)
            
        return result
        
    def process_all_vouchers(self, project: Optional[str] = None) -> List[ProcessingResult]:
        """
        处理所有凭证文件
        
        Args:
            project: 指定项目名称，None表示处理所有项目
            
        Returns:
            处理结果列表
        """
        results = []
        
        mapper = ProjectFolderMapper(self.project_detector, self.config.get('paths', {}))
        voucher_files = mapper.scan_all_voucher_files()
        
        total_files = sum(len(files) for files in voucher_files.values())
        processed = 0
        
        for proj_name, files in voucher_files.items():
            if project and proj_name != project:
                continue
                
            logger.info(f"Processing project: {proj_name} ({len(files)} files)")
            
            for pdf_file in files:
                processed += 1
                logger.info(f"[{processed}/{total_files}] Processing: {pdf_file.name}")
                
                result = self.process_voucher(str(pdf_file))
                results.append(result)
                
        return results


class ReportGenerator:
    """报告生成器"""
    
    def __init__(self, output_dir: Path):
        """
        初始化报告生成器
        
        Args:
            output_dir: 输出目录
        """
        self.output_dir = output_dir
        
    def generate_match_report(self, results: List[ProcessingResult], 
                              project: str) -> str:
        """
        生成项目匹配报告
        
        Args:
            results: 处理结果列表
            project: 项目名称
            
        Returns:
            报告文件路径
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("openpyxl not installed. Please run: pip install openpyxl")
            return ""
            
        # 创建工作簿
        wb = openpyxl.Workbook()
        
        # Sheet 1: 凭证匹配结果
        ws1 = wb.active
        ws1.title = "凭证匹配结果"
        
        # 表头
        headers = ["凭证文件名", "凭证页码", "文档类型", "匹配状态", 
                   "匹配PDF", "匹配页码", "相似度", "关键匹配词"]
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            
        # 数据行
        row = 2
        for result in results:
            if result.project != project:
                continue
                
            for match in result.matches:
                ws1.cell(row=row, column=1, value=Path(match.source_file).name)
                ws1.cell(row=row, column=2, value=match.source_pages)
                ws1.cell(row=row, column=3, value=match.doc_type)
                ws1.cell(row=row, column=4, value=match.match_status)
                ws1.cell(row=row, column=5, value=Path(match.target_file).name if match.target_file else "-")
                ws1.cell(row=row, column=6, value=match.target_pages if match.target_pages else "-")
                ws1.cell(row=row, column=7, value=f"{match.similarity:.0%}" if match.similarity else "-")
                ws1.cell(row=row, column=8, value="、".join(match.matched_keywords) if match.matched_keywords else "-")
                
                # 根据匹配状态设置颜色
                status_cell = ws1.cell(row=row, column=4)
                if match.match_status == "完全匹配":
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif match.match_status == "部分匹配":
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif match.match_status == "未找到":
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                row += 1
                
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
        # Sheet 2: 统计汇总
        ws2 = wb.create_sheet("统计汇总")
        
        # 统计数据
        project_results = [r for r in results if r.project == project]
        all_matches = [m for r in project_results for m in r.matches]
        
        stats = [
            ("项目名称", project),
            ("处理文件数", len(project_results)),
            ("总页面数", sum(len(r.pages) for r in project_results)),
            ("完全匹配", sum(1 for m in all_matches if m.match_status == "完全匹配")),
            ("部分匹配", sum(1 for m in all_matches if m.match_status == "部分匹配")),
            ("未找到", sum(1 for m in all_matches if m.match_status == "未找到")),
            ("匹配率", f"{sum(1 for m in all_matches if m.match_status != '未找到') / len(all_matches):.1%}" if all_matches else "N/A"),
        ]
        
        for row, (label, value) in enumerate(stats, 1):
            ws2.cell(row=row, column=1, value=label)
            ws2.cell(row=row, column=2, value=value)
            
        # 保存报告
        project_dir = self.output_dir / project
        project_dir.mkdir(parents=True, exist_ok=True)
        
        report_path = project_dir / "匹配报告.xlsx"
        wb.save(str(report_path))
        
        logger.info(f"Report saved: {report_path}")
        return str(report_path)
        
    def generate_summary_report(self, results: List[ProcessingResult]) -> str:
        """
        生成汇总报告（增强版）
        
        Args:
            results: 所有处理结果
            
        Returns:
            报告文件路径
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.chart import BarChart, Reference, PieChart
        except ImportError:
            logger.error("openpyxl not installed")
            return ""
            
        wb = openpyxl.Workbook()
        
        # Sheet 1: 汇总报告
        self._create_summary_sheet(wb, results)
        
        # Sheet 2: 处理统计
        self._create_processing_stats_sheet(wb, results)
        
        # Sheet 3: 识别质量分析
        self._create_quality_analysis_sheet(wb, results)
        
        # Sheet 4: 关键词统计
        self._create_keyword_analysis_sheet(wb, results)
        
        # Sheet 5: 时间线视图
        self._create_timeline_sheet(wb, results)
        
        # 保存
        report_path = self.output_dir / "汇总报告.xlsx"
        wb.save(str(report_path))
        
        logger.info(f"Enhanced summary report saved: {report_path}")
        return str(report_path)
    
    def _create_summary_sheet(self, wb, results):
        """创建汇总Sheet"""
        from openpyxl.styles import Font, PatternFill
        
        ws = wb.active
        ws.title = "汇总报告"
        
        headers = ["项目名称", "文件数", "总页数", "完全匹配", "部分匹配", "未找到", "匹配率"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
        # 按项目统计
        projects = {}
        for result in results:
            if result.project not in projects:
                projects[result.project] = {'files': 0, 'pages': 0, 'exact': 0, 'partial': 0, 'not_found': 0}
            projects[result.project]['files'] += 1
            projects[result.project]['pages'] += len(result.pages)
            for match in result.matches:
                if match.match_status == "完全匹配":
                    projects[result.project]['exact'] += 1
                elif match.match_status == "部分匹配":
                    projects[result.project]['partial'] += 1
                else:
                    projects[result.project]['not_found'] += 1
                    
        row = 2
        for project, stats in sorted(projects.items()):
            total_matches = stats['exact'] + stats['partial'] + stats['not_found']
            match_rate = (stats['exact'] + stats['partial']) / total_matches if total_matches > 0 else 0
            
            ws.cell(row=row, column=1, value=project)
            ws.cell(row=row, column=2, value=stats['files'])
            ws.cell(row=row, column=3, value=stats['pages'])
            ws.cell(row=row, column=4, value=stats['exact'])
            ws.cell(row=row, column=5, value=stats['partial'])
            ws.cell(row=row, column=6, value=stats['not_found'])
            ws.cell(row=row, column=7, value=f"{match_rate:.1%}")
            row += 1
    
    def _create_processing_stats_sheet(self, wb, results):
        """创建处理统计Sheet"""
        from openpyxl.styles import Font, PatternFill
        
        ws = wb.create_sheet("处理统计")
        
        headers = ["文件名", "项目", "页数", "处理时间(s)", "速度(s/页)", "OCR引擎", "状态"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        row = 2
        total_time = 0
        total_pages = 0
        
        for result in results:
            filename = Path(result.file_path).name
            pages = len(result.pages)
            proc_time = getattr(result, 'processing_time', 0)
            speed = proc_time / pages if pages > 0 else 0
            engine = getattr(result, 'engine_used', 'unknown')
            status = "✅ 成功" if result.success else "❌ 失败"
            
            ws.cell(row=row, column=1, value=filename)
            ws.cell(row=row, column=2, value=result.project)
            ws.cell(row=row, column=3, value=pages)
            ws.cell(row=row, column=4, value=f"{proc_time:.1f}")
            ws.cell(row=row, column=5, value=f"{speed:.2f}")
            ws.cell(row=row, column=6, value=engine)
            ws.cell(row=row, column=7, value=status)
            
            total_time += proc_time
            total_pages += pages
            row += 1
        
        # 汇总行
        ws.cell(row=row, column=1, value="合计")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_pages)
        ws.cell(row=row, column=4, value=f"{total_time:.1f}")
        avg_speed = total_time / total_pages if total_pages > 0 else 0
        ws.cell(row=row, column=5, value=f"{avg_speed:.2f}")
    
    def _create_quality_analysis_sheet(self, wb, results):
        """创建识别质量分析Sheet"""
        from openpyxl.styles import Font, PatternFill
        
        ws = wb.create_sheet("识别质量")
        
        headers = ["文件名", "页码", "平均置信度", "最低置信度", "文本长度", "质量评级"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        row = 2
        for result in results:
            for page in result.pages:
                page_num = page.get('page_num', 0)
                avg_conf = page.get('avg_confidence', 1.0)
                min_conf = page.get('min_confidence', 1.0)
                text_len = len(page.get('text', ''))
                
                # 质量评级
                if avg_conf >= 0.9:
                    quality = "🟢 优"
                elif avg_conf >= 0.8:
                    quality = "🟡 良"
                elif avg_conf >= 0.6:
                    quality = "🟠 中"
                else:
                    quality = "🔴 差"
                
                ws.cell(row=row, column=1, value=Path(result.file_path).name)
                ws.cell(row=row, column=2, value=page_num)
                ws.cell(row=row, column=3, value=f"{avg_conf:.2f}")
                ws.cell(row=row, column=4, value=f"{min_conf:.2f}")
                ws.cell(row=row, column=5, value=text_len)
                ws.cell(row=row, column=6, value=quality)
                row += 1
    
    def _create_keyword_analysis_sheet(self, wb, results):
        """创建关键词统计Sheet"""
        from openpyxl.styles import Font, PatternFill
        from collections import Counter
        
        ws = wb.create_sheet("关键词分析")
        
        headers = ["关键词", "出现次数", "出现文件数", "相关项目"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        
        # 统计关键词
        keyword_counter = Counter()
        keyword_files = {}
        keyword_projects = {}
        
        for result in results:
            for page in result.pages:
                keywords = page.get('keywords', [])
                for kw in keywords:
                    keyword_counter[kw] += 1
                    if kw not in keyword_files:
                        keyword_files[kw] = set()
                        keyword_projects[kw] = set()
                    keyword_files[kw].add(Path(result.file_path).name)
                    keyword_projects[kw].add(result.project)
        
        row = 2
        for kw, count in keyword_counter.most_common(50):  # Top 50
            ws.cell(row=row, column=1, value=kw)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=len(keyword_files.get(kw, set())))
            ws.cell(row=row, column=4, value="、".join(keyword_projects.get(kw, set())))
            row += 1
    
    def _create_timeline_sheet(self, wb, results):
        """创建时间线视图Sheet"""
        from openpyxl.styles import Font, PatternFill
        from collections import defaultdict
        
        ws = wb.create_sheet("时间线")
        
        headers = ["日期", "文件数", "页数", "项目", "文件列表"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        
        # 按日期分组
        date_groups = defaultdict(list)
        
        for result in results:
            for page in result.pages:
                dates = page.get('dates', [])
                for date in dates:
                    date_groups[date].append({
                        'file': Path(result.file_path).name,
                        'project': result.project,
                        'pages': len(result.pages)
                    })
        
        row = 2
        for date in sorted(date_groups.keys()):
            items = date_groups[date]
            files = set(item['file'] for item in items)
            projects = set(item['project'] for item in items)
            total_pages = sum(item['pages'] for item in items)
            
            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=2, value=len(files))
            ws.cell(row=row, column=3, value=total_pages)
            ws.cell(row=row, column=4, value="、".join(projects))
            ws.cell(row=row, column=5, value="、".join(list(files)[:3]) + ("..." if len(files) > 3 else ""))
            row += 1


def load_config(config_path: str = None) -> dict:
    """加载配置文件"""
    if config_path is None:
        config_path = Path(__file__).parent / "config.yaml"
        
    if not os.path.exists(config_path):
        logger.error(f"Config file not found: {config_path}")
        raise FileNotFoundError(f"Config file not found: {config_path}")
        
    with open(config_path, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
        
    logger.info(f"Config loaded from: {config_path}")
    return config


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description="PDF智能OCR分类与比对系统",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument(
        '--config', '-c',
        type=str,
        default=None,
        help='配置文件路径 (默认: config.yaml)'
    )
    
    parser.add_argument(
        '--project', '-p',
        type=str,
        default=None,
        help='指定处理的项目名称 (默认: 处理所有项目)'
    )
    
    parser.add_argument(
        '--input', '-i',
        type=str,
        default=None,
        help='指定输入文件或目录 (覆盖配置文件设置)'
    )
    
    parser.add_argument(
        '--ocr-only',
        action='store_true',
        help='仅执行OCR识别，不进行比对'
    )
    
    parser.add_argument(
        '--match-only',
        action='store_true',
        help='仅执行比对，假设OCR结果已存在'
    )
    
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='试运行模式，不实际处理文件'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='显示详细日志'
    )
    
    parser.add_argument(
        '--clear-cache',
        action='store_true',
        help='清除OCR缓存后退出'
    )
    
    parser.add_argument(
        '--index-only',
        action='store_true',
        help='只构建参照资料索引，不处理凭证'
    )
    
    args = parser.parse_args()
    
    # 设置日志级别
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
        
    try:
        # 加载配置
        config = load_config(args.config)
        
        # 覆盖输入路径
        if args.input:
            config['paths']['input_vouchers'] = args.input
            
        # 试运行模式
        if args.dry_run:
            logger.info("=== DRY RUN MODE ===")
            logger.info(f"Config: {config['paths']}")
            logger.info(f"Project filter: {args.project or 'All'}")
            return
        
        # 清除缓存模式
        if args.clear_cache:
            from ocr_cache import OCRCache
            cache_dir = Path(config.get('paths', {}).get('temp_dir', './temp')) / 'ocr_cache'
            cache = OCRCache(str(cache_dir))
            cache.clear()
            logger.info("OCR cache cleared successfully")
            return
            
        # 初始化处理器
        processor = DocumentProcessor(config)
        
        # 构建参照资料索引
        if not args.ocr_only:
            processor.build_reference_index(args.project)
        
        # 如果只是构建索引，则退出
        if args.index_only:
            logger.info("Index-only mode complete")
            return
            
        # 处理凭证文件
        if not args.match_only:
            results = processor.process_all_vouchers(args.project)
            
            # 生成报告
            report_generator = ReportGenerator(processor.output_path)
            
            # 按项目生成匹配报告
            projects = set(r.project for r in results)
            for project in projects:
                report_generator.generate_match_report(results, project)
                
            # 生成汇总报告
            report_generator.generate_summary_report(results)
            
            # 输出统计
            success_count = sum(1 for r in results if r.success)
            logger.info(f"\n=== Processing Complete ===")
            logger.info(f"Total files: {len(results)}")
            logger.info(f"Successful: {success_count}")
            logger.info(f"Failed: {len(results) - success_count}")
            logger.info(f"Output directory: {processor.output_path}")
            
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
